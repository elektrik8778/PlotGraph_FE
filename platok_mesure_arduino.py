import serial
import datetime
import openpyxl
from openpyxl import load_workbook

mywb = openpyxl.Workbook()


sheet_name = str(datetime.datetime.now().time()).split(".")[0].replace(':',"-")
print(sheet_name)
mywb.create_sheet(index=0, title=f'{sheet_name}')
print(mywb.sheetnames)

activ_sheet = mywb[f'{sheet_name}']
activ_sheet.column_dimensions['A'].width = 17.5
activ_sheet.column_dimensions['B'].width = 13
activ_sheet[f'A1'] = f'Время измерения'
activ_sheet[f'B1'] = f'Температура'


for i in range(2,11):
    activ_sheet[f'A{i}'] = str(datetime.datetime.now().time()).split(".")[0]
    activ_sheet[f'B{i}'] = 23 + i


mywb.save('example.xlsx')

# with serial.Serial('COM4', 9600, timeout=1) as ser:
#     while True:
#         line = ser.readline()
#         if line:
#             data = line.strip()
#             data = str(data.decode())
#             if data[0] == 'T':
#                 print(f'{str(datetime.datetime.now().time()).split(".")[0]} Температура: {data[1:]} C')
